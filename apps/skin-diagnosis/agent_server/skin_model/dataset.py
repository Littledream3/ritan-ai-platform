import os
import torch
from torch.utils.data import Dataset
from PIL import Image
from openpyxl import load_workbook

METRIC_NAMES = [
    'acne_severity', 'blackheads', 'whiteheads', 'open_pores',
    'excessive_oil', 'skin_irritation', 'skin_sensitivity', 'redness',
    'fine_lines_eyes', 'eye_puffiness', 'dark_circles', 'forehead_wrinkles',
    'skin_elasticity', 'dehydration', 'dark_spots',
    'post_acne_marks', 'uneven_skin', 'freckles',
]


class SkinMetricsDataset(Dataset):
    def __init__(self, image_paths, labels, transform=None):
        self.image_paths = image_paths
        self.labels = labels  # FloatTensor (N, 18)
        self.transform = transform

    def __len__(self):
        return len(self.image_paths)

    def __getitem__(self, idx):
        img = Image.open(self.image_paths[idx]).convert('RGB')
        if self.transform:
            img = self.transform(img)
        return img, self.labels[idx]


_KNOWN_EXTS = ('.jpg', '.jpeg', '.png', '.JPG', '.JPEG', '.PNG')

def _strip_ext(name: str) -> str:
    """Remove image extension if present, leaving the full stem intact."""
    name = name.strip()
    for ext in _KNOWN_EXTS:
        if name.endswith(ext):
            return name[: -len(ext)]
    return name


def load_excel_data(excel_dir, image_dir):
    """
    Read both Excel files, match Image_ID to files in image_dir,
    return (list[str], FloatTensor[N,18]).
    """
    # Build lookup: stem (no extension) -> full_path
    file_map = {}
    for cls in ['oily', 'dry', 'normal']:
        folder = os.path.join(image_dir, cls)
        if not os.path.exists(folder):
            continue
        for fname in os.listdir(folder):
            file_map[_strip_ext(fname)] = os.path.join(folder, fname)

    all_paths, all_labels = [], []
    seen_stems = set()  # deduplicate across the two Excel files

    for excel_name in [
        'skinalaysis_labeling_train1.xlsx',
        'skinanalysis_valid1.xlsx',
    ]:
        wb = load_workbook(os.path.join(excel_dir, excel_name))
        for row in wb.active.iter_rows(values_only=True, min_row=2):
            img_id = row[1]
            if not img_id:
                continue
            stem = _strip_ext(str(img_id))
            if stem in seen_stems or stem not in file_map:
                continue
            seen_stems.add(stem)

            # Columns 2-19 are the 18 metrics; clamp to [0, 5]
            metrics = []
            for i in range(2, 20):
                v = row[i] if row[i] is not None else 0.0
                metrics.append(float(max(0.0, min(5.0, v))))

            all_paths.append(file_map[stem])
            all_labels.append(metrics)

    return all_paths, torch.tensor(all_labels, dtype=torch.float32)
