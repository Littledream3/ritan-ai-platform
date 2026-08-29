"""飞书事件回调 API — 接收飞书消息 → 调用模型 → 回复（支持文本/文件/图片）"""

import base64
import io
import json
import os
import re
import time
from datetime import datetime

from fastapi import APIRouter, BackgroundTasks, HTTPException, Request
from loguru import logger

from app.adapters.registry import registry
from app.services.search_service import search_web, is_weather_query, get_weather
from app.services import media_service as media
from app.services.feishu_service import (
    decrypt_event,
    download_resource,
    fetch_recent_file,
    get_session_model,
    is_event_processed,
    reply_file_message,
    reply_image_message,
    reply_text_message,
    send_text_to_user,
    set_session_model,
    upload_file,
    upload_image,
)

router = APIRouter(prefix="/api/v1/feishu", tags=["飞书"])

# 模型显示名映射
MODEL_LABELS = {
    "deepseek-v4-pro": "DeepSeek V4 Pro",
    "kimi-k2.7-code": "Kimi K2.7 Code",
    "qwen3.7-max": "Qwen 3.7 Max",
    "glm-4.7": "GLM 4.7",
    "doubao-seed-2-0-lite-260215": "豆包 Seed 2.0 Lite（免费版）",
}

DEFAULT_MODEL = "kimi-k2.7-code"

# 视觉模型 fallback 顺序
VISION_MODELS = ["kimi-k2.7-code", "doubao-seed-2-0-lite-260215"]

# 支持的文件后缀
SUPPORTED_EXTENSIONS = {
    ".txt", ".md", ".py", ".js", ".ts", ".html", ".css", ".json", ".xml",
    ".csv", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".conf",
    ".pdf", ".docx", ".xlsx",
}


# ---- 事件回调入口 ----


@router.post("/event")
async def feishu_event(request: Request, background: BackgroundTasks):
    """飞书事件回调端点"""
    body = await request.body()
    body_str = body.decode("utf-8")

    try:
        data = json.loads(body_str)
    except json.JSONDecodeError:
        logger.warning("飞书回调: 非 JSON 请求体")
        raise HTTPException(status_code=400, detail="Invalid JSON")

    # URL 验证
    if data.get("type") == "url_verification":
        challenge = data.get("challenge", "")
        logger.info(f"飞书 URL 验证: challenge={challenge[:20]}...")
        return {"challenge": challenge}

    # 加密事件
    if "encrypt" in data:
        try:
            data = decrypt_event(data["encrypt"])
        except Exception as e:
            logger.error(f"飞书事件解密失败: {e}")
            raise HTTPException(status_code=400, detail="Decrypt failed")
        if data.get("type") == "url_verification":
            return {"challenge": data.get("challenge", "")}

    # 新版事件
    header = data.get("header", {})
    event_type = header.get("event_type", "")
    event_id = header.get("event_id", "")
    event_data = data.get("event", {})

    logger.info(f"飞书事件: type={event_type}, event_id={event_id}")

    if event_id and await is_event_processed(event_id):
        logger.info(f"事件 {event_id} 已处理，跳过")
        return {"code": 0}

    if event_type == "im.message.receive_v1":
        msg = event_data.get("message", {})
        logger.info(f"消息详情: msg_type={msg.get('message_type')}, msg_id={msg.get('message_id')}")
        background.add_task(_handle_message, event_data)
    else:
        logger.info(f"忽略事件类型: {event_type}")

    return {"code": 0}


# ---- 消息处理 ----


def _is_at_bot(text: str, mentions: list = None) -> bool:
    """检查消息是否 @了机器人

    飞书 @提及有两种方式：
    1. mentions 数组非空 → 有人被 @
    2. 文本中有 <at user_id="xxx">name</at> 标签
    """
    if mentions:
        return True
    return "<at " in text or "@" in text


async def _handle_message(event: dict):
    """后台处理消息: 解析 → 路由 → 回复"""
    try:
        message = event.get("message", {})
        sender = event.get("sender", {})
        sender_id = sender.get("sender_id", {})
        open_id = sender_id.get("open_id", "")
        message_id = message.get("message_id", "")
        chat_type = message.get("chat_type", "p2p")
        chat_id = message.get("chat_id", "")
        msg_type = message.get("message_type", "")

        logger.info(f"飞书消息: type={msg_type}, open_id={open_id}, chat={chat_type}, chat_id={chat_id}")

        if not open_id:
            logger.warning("飞书消息缺少 sender.open_id")
            return

        # 群聊中只有 @机器人 才回复，私聊始终回复
        is_group = chat_type == "group"
        mentions = message.get("mentions", []) or []
        logger.info(f"mentions={mentions}, is_group={is_group}")

        # ---- 图片消息 ----
        if msg_type == "image":
            if is_group:
                logger.info("群聊图片消息，跳过（需 @机器人）")
                return
            response_text = await _handle_image(open_id, message_id, message)
        # ---- 文件消息 ----
        elif msg_type == "file":
            if is_group:
                logger.info("群聊文件消息，跳过（请用 /read 命令读取）")
                return
            response_text = await _handle_file(open_id, message_id, message)
        # ---- 富文本（可能含图片）----
        elif msg_type == "post":
            raw_content = message.get("content", "{}")
            logger.info(f"POST 原始内容: {raw_content[:500]}")
            result = _parse_post_content(raw_content)
            logger.info(f"解析结果: text={result['text'][:100]}, images={result['image_keys']}")
            if result["image_keys"] and message_id:
                if is_group and not _is_at_bot(raw_content, mentions):
                    logger.info(f"群聊 post 图片消息未 @机器人，跳过 (raw={raw_content[:200]})")
                    return
                response_text = await _handle_post_image(open_id, message_id, result)
            elif result["text"]:
                if is_group and not _is_at_bot(result["text"], mentions):
                    logger.info(f"群聊 post 消息未 @机器人，跳过 (text={result['text'][:200]})")
                    return
                text = _clean_at_mentions(result["text"])
                logger.info(f"飞书消息: open_id={open_id}, text={text[:100]}, chat_type={chat_type}")
                response_text = await _route_message(open_id, text, chat_id, message_id)
            else:
                return
        # ---- 纯文本 ----
        else:
            text = ""
            if msg_type == "text":
                content_str = message.get("content", "{}")
                try:
                    text = json.loads(content_str).get("text", "").strip()
                except json.JSONDecodeError:
                    text = content_str

            if not text:
                return

            # 群聊中必须 @机器人 才回复
            if is_group and not _is_at_bot(text, mentions):
                logger.info(f"群聊文本消息未 @机器人，跳过 (raw_text={text[:200]}, mentions={mentions})")
                return

            text = _clean_at_mentions(text)
            logger.info(f"飞书消息: open_id={open_id}, text={text[:100]}, chat_type={chat_type}")
            response_text = await _route_message(open_id, text, chat_id, message_id)

        # 发送回复
        if message_id:
            result = await reply_text_message(message_id, response_text)
            if result.get("code") != 0:
                logger.error(f"回复消息失败: {result}")
                if chat_type == "p2p":
                    await send_text_to_user(open_id, response_text)
        else:
            await send_text_to_user(open_id, response_text)

    except Exception as e:
        logger.error(f"处理飞书消息异常: {e}", exc_info=True)


# ---- 图片处理 ----

def _get_vision_adapter(open_id: str, preferred_model: str):
    """
    获取视觉识别用的适配器，自动 fallback

    如果当前模型不支持 vision（如 DeepSeek、GLM），
    自动切换到第一个可用的视觉模型（Kimi）。

    Returns:
        (model_id, adapter, label, is_fallback)
    """
    # 检查首选模型是否支持 vision
    try:
        adapter = registry.get(preferred_model)
        if "vision" in adapter.model_info.capabilities and adapter.model_info.available:
            label = MODEL_LABELS.get(preferred_model, preferred_model)
            return preferred_model, adapter, label, False
    except ValueError:
        pass

    # fallback 到第一个视觉模型
    for vm in VISION_MODELS:
        try:
            adapter = registry.get(vm)
            if adapter.model_info.available:
                label = MODEL_LABELS.get(vm, vm)
                return vm, adapter, label, True
        except ValueError:
            continue

    # 最后兜底 — 用默认模型（可能不支持 vision，但至少不报错）
    adapter = registry.get(DEFAULT_MODEL)
    label = MODEL_LABELS.get(DEFAULT_MODEL, DEFAULT_MODEL)
    return DEFAULT_MODEL, adapter, label, True


async def _handle_image(open_id: str, message_id: str, message: dict) -> str:
    """处理图片消息: 下载 → base64 → 交给模型"""
    content = json.loads(message.get("content", "{}"))
    image_key = content.get("image_key", "")

    if not image_key or not message_id:
        return "无法读取图片，请重试。"

    try:
        image_bytes = await download_resource(message_id, image_key, "image")
    except Exception as e:
        logger.error(f"下载图片失败: {e}")
        return "图片下载失败，文件可能过大。"

    img_b64 = base64.b64encode(image_bytes).decode()
    # 检测图片格式
    ext = ".png" if image_bytes[:8] == b'\x89PNG\r\n\x1a\n' else ".jpg"
    mime = "image/png" if ext == ".png" else "image/jpeg"

    vision_content = [
        {"type": "text", "text": "请描述这张图片的内容。"},
        {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{img_b64}"}},
    ]

    preferred = await get_session_model(open_id)
    original_label = MODEL_LABELS.get(preferred, preferred)

    model, adapter, label, is_fallback = _get_vision_adapter(open_id, preferred)

    try:
        result = await adapter.chat(messages=[{"role": "user", "content": vision_content}])
    except Exception as e:
        logger.error(f"图片识别失败: model={model}, error={e}")
        return f"图片识别失败，请稍后重试。\n错误: {str(e)[:200]}"

    content_text = ""
    choices = result.get("choices", [])
    if choices:
        content_text = choices[0].get("message", {}).get("content", "")

    if not content_text:
        return "模型未能识别图片内容。"

    footnote = f"📷 图片识别 · 模型: **{label}**"
    if is_fallback:
        footnote += f"\n💡 原模型 **{original_label}** 不支持图片，已自动切换"
    return f"{content_text}\n\n———\n{footnote}"


# ---- 文件处理 ----


async def _handle_file(open_id: str, message_id: str, message: dict) -> str:
    """处理文件消息: 下载 → 提取文本 → 交给模型"""
    content = json.loads(message.get("content", "{}"))
    file_key = content.get("file_key", "")
    # 文件名可能在多个位置: content JSON、body.title、或需从 API 获取
    file_name = content.get("file_name") or message.get("file_name") or ""
    logger.info(f"文件消息: file_key={file_key[:30]}..., file_name_from_content={content.get('file_name')}, file_name_from_msg={message.get('file_name')}")
    if not file_name:
        from app.services.feishu_service import get_message_info
        try:
            info = await get_message_info(message_id)
            body = info.get("data", {}).get("items", [{}])[0].get("body", {})
            body_content = json.loads(body.get("content", "{}"))
            file_name = body_content.get("file_name") or body.get("title", "unknown")
            logger.info(f"从 API 获取文件名: {file_name}")
        except Exception as e:
            logger.warning(f"获取文件名失败: {e}")
            file_name = "unknown"

    if not file_key or not message_id:
        return "无法读取文件，请重试。"

    file_name_lower = file_name.lower()
    ext = file_name_lower[file_name_lower.rfind("."):] if "." in file_name_lower else ""

    if ext not in SUPPORTED_EXTENSIONS:
        return f"暂不支持 {ext} 格式的文件。\n支持: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"

    # 先回复一条"处理中"
    logger.info(f"处理文件: {file_name} (key={file_key})")

    try:
        file_bytes = await download_resource(message_id, file_key, "file")
    except Exception as e:
        logger.error(f"下载文件失败: {e}")
        return f"文件下载失败。\n{str(e)[:200]}"

    # 提取文本
    try:
        file_text = _extract_file_text(file_bytes, ext)
    except Exception as e:
        logger.error(f"提取文件文本失败: {e}")
        return f"文件解析失败: {str(e)[:200]}"

    if not file_text:
        return "文件为空或无法提取文本内容。"

    # 截断过长文本
    max_chars = 12000
    if len(file_text) > max_chars:
        file_text = file_text[:max_chars] + f"\n\n...(文件过大，已截断，剩余 {len(file_text) - max_chars} 字符)"
    logger.info(f"提取文本 {len(file_text)} 字符，准备调用模型...")

    model = await get_session_model(open_id)
    label = MODEL_LABELS.get(model, model)

    format_instruction = _get_format_instruction()
    prompt = f"请分析以下文件内容（文件: {file_name}）：\n\n```\n{file_text}\n```\n\n请总结这个文件的核心内容。{format_instruction}"

    try:
        adapter = registry.get(model)
    except ValueError:
        adapter = registry.get(DEFAULT_MODEL)
        model = DEFAULT_MODEL
        label = MODEL_LABELS.get(model, model)

    try:
        result = await adapter.chat(messages=[{"role": "user", "content": prompt}])
    except Exception as e:
        logger.error(f"文件分析失败: type={type(e).__name__}, msg={e}")
        return f"文件分析失败: [{type(e).__name__}] {str(e)[:200] or '未知错误'}"

    content_text = ""
    choices = result.get("choices", [])
    if choices:
        content_text = choices[0].get("message", {}).get("content", "")

    if not content_text:
        return "模型未能分析文件内容。"

    return f"{content_text}\n\n———\n📁 **{file_name}** · 模型: **{label}**"


def _extract_file_text(file_bytes: bytes, ext: str) -> str:
    """从文件二进制内容中提取文本"""
    if ext in (".txt", ".md", ".py", ".js", ".ts", ".html", ".css",
               ".json", ".xml", ".csv", ".yaml", ".yml", ".toml",
               ".ini", ".cfg", ".conf"):
        return file_bytes.decode("utf-8", errors="replace")

    if ext == ".pdf":
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(file_bytes))
        return "\n\n".join(page.extract_text() or "" for page in reader.pages)

    if ext == ".docx":
        from docx import Document
        doc = Document(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs)

    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        parts = []
        for name in wb.sheetnames:
            ws = wb[name]
            parts.append(f"=== {name} ===")
            for row in ws.iter_rows(values_only=True):
                parts.append("\t".join(str(c) if c is not None else "" for c in row))
        return "\n".join(parts)

    return ""


# ---- 工具函数 ----


def _get_format_instruction() -> str:
    """飞书聊天格式化指令 — 让模型输出适合飞书阅读的纯文本"""
    return (
        "\n\n⚠️ 注意：你的回复将显示在飞书聊天中，请用清晰易读的纯文本格式输出。"
        "数学公式用 ASCII 方式书写（如 sqrt(x) 代替 \\sqrt{x}，x^2 用 x² 或 x^2，"
        "α/β 等希腊字母直接写 alpha/beta），合理使用空行分段、编号列表和分隔线。"
    )


def _clean_at_mentions(text: str) -> str:
    """移除 @机器人 等提及标记"""
    text = re.sub(r'@\S+\s*', '', text)
    return text.strip()


def _parse_post_content(content_str: str) -> dict:
    """
    从富文本 post 中提取文本和图片

    飞书 post 格式:
      {"title":"", "content": [[{tag,text,image_key,...}, ...], ...]}
      或
      {"zh_cn": {"title":"", "content": [...]}}

    Returns:
        {"text": "...", "image_keys": ["img_v3_xxx", ...]}
    """
    result = {"text": "", "image_keys": []}
    try:
        post = json.loads(content_str)

        # 获取 content 数组 — 兼容有无语言包两种格式
        content_list = []
        if "content" in post and isinstance(post["content"], list):
            content_list = post["content"]
        else:
            for lang, body in post.items():
                if isinstance(body, dict) and "content" in body:
                    content_list = body.get("content", [])
                    break

        # 遍历 content (嵌套列表: 段落 → 元素)
        text_parts = []
        for para in content_list:
            if not isinstance(para, list):
                continue
            for elem in para:
                if not isinstance(elem, dict):
                    continue
                tag = elem.get("tag", "")
                if tag == "text":
                    text_parts.append(str(elem.get("text", "")))
                elif tag == "img":
                    ik = elem.get("image_key", "")
                    if ik:
                        result["image_keys"].append(ik)
                elif tag == "at":
                    pass  # 跳过 @提及

        result["text"] = "".join(text_parts)
    except Exception as e:
        logger.error(f"解析 post 内容失败: {e}")
    return result


async def _handle_post_image(open_id: str, message_id: str, parsed: dict) -> str:
    """处理富文本中夹带的图片（支持自动 vision fallback）"""
    image_key = parsed["image_keys"][0]
    user_text = _clean_at_mentions(parsed["text"]).strip()

    try:
        image_bytes = await download_resource(message_id, image_key, "image")
    except Exception as e:
        logger.error(f"下载图片失败: {e}")
        return "图片下载失败，请重试。"

    img_b64 = base64.b64encode(image_bytes).decode()
    ext = ".png" if image_bytes[:8] == b'\x89PNG\r\n\x1a\n' else ".jpg"
    mime = "image/png" if ext == ".png" else "image/jpeg"

    prompt = user_text if user_text else "请描述这张图片的内容。"
    vision_content = [
        {"type": "text", "text": prompt},
        {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{img_b64}"}},
    ]

    preferred = await get_session_model(open_id)
    original_label = MODEL_LABELS.get(preferred, preferred)

    model, adapter, label, is_fallback = _get_vision_adapter(open_id, preferred)

    try:
        result = await adapter.chat(messages=[{"role": "user", "content": vision_content}])
    except Exception as e:
        logger.error(f"图片识别失败: {e}")
        return f"图片识别失败。\n{str(e)[:200]}"

    content_text = ""
    choices = result.get("choices", [])
    if choices:
        content_text = choices[0].get("message", {}).get("content", "")

    if not content_text:
        return "模型未能识别图片内容。"

    footnote = f"📷 图片识别 · 模型: **{label}**"
    if is_fallback:
        footnote += f"\n💡 原模型 **{original_label}** 不支持图片，已自动切换"
    return f"{content_text}\n\n———\n{footnote}"


def _extract_text_from_post(content_str: str) -> str:
    """从富文本 post 中提取纯文本（向后兼容）"""
    return _parse_post_content(content_str)["text"]


# ---- 消息路由 ----


async def _route_message(open_id: str, text: str, chat_id: str = "", message_id: str = "") -> str:
    """根据消息内容路由"""
    if not text:
        return "请发送消息或指令。输入 /help 查看帮助。"

    if text == "/help":
        return _cmd_help()
    if text == "/models":
        return _cmd_models()
    if text.startswith("/model"):
        return await _cmd_model(open_id, text)
    if text.startswith("/search") or text.startswith("搜索") or text.startswith("查一下"):
        parts = text.split(maxsplit=1)
        query = parts[1].strip() if len(parts) > 1 else ""
        if not query:
            return "🔍 **联网搜索**\n\n用法：`/search 关键词`\n例如：`/search 今天有什么新闻`"
        return await _cmd_search(open_id, query, message_id)
    if any(text.startswith(cmd) for cmd in ("/clip", "/compress", "/watermark", "/img-watermark",
            "/extract-audio", "/convert", "/resize", "/img-compress", "/concat", "/info")):
        return await _cmd_media(open_id, chat_id, message_id, text)
    if text.startswith("/read") or text.startswith("读文件") or text.startswith("读取文件") or text.startswith("分析文件") or text.startswith("看看文件"):
        # 支持: /read | /read report.pdf | /read report.pdf 总结一下
        #        /read report.pdf，总结一下（中文逗号也支持）
        parts = text.split(maxsplit=1)
        args = parts[1].strip() if len(parts) > 1 else ""
        # 按空格/中文逗号/英文逗号拆分文件名和自定义问题
        arg_parts = re.split(r'[ ，,]\s*', args, maxsplit=1)
        filename = arg_parts[0].strip() if arg_parts else ""
        question = arg_parts[1].strip() if len(arg_parts) > 1 else ""
        if filename:
            return await _cmd_read(open_id, chat_id, filename, question)
        return await _cmd_read(open_id, chat_id, "", "")

    return await _cmd_chat(open_id, text)


# ---- 指令实现 ----


def _cmd_help() -> str:
    return (
        "🤖 **多模型 AI 助手**\n\n"
        "💬 **对话 & 搜索**\n"
        "• `/help` — 显示此帮助\n"
        "• `/models` — 查看可用模型\n"
        "• `/model <名称>` — 切换模型\n"
        "• `/search <关键词>` — 联网搜索\n"
        "• `/read` — 分析群聊中最近的文件\n\n"
        "🎬 **视频处理**（先发视频再用命令）\n"
        "• `/clip 00:05 00:30` — 裁剪片段\n"
        "• `/compress [crf]` — 压缩视频（如 /compress 28）\n"
        "• `/watermark <文字> [位置] [字号]` — 加水印\n"
        "• `/extract-audio [格式]` — 提取音频\n"
        "• `/convert <格式>` — 格式转换（如 /convert gif）\n"
        "• `/concat` — 拼接多个视频\n"
        "• `/info` — 查看视频信息\n\n"
        "🖼 **图片处理**（先发图片再用命令）\n"
        "• `/resize 800` — 按宽度缩放\n"
        "• `/img-watermark <文字> [位置] [字号]` — 图片加水印\n"
        "• `/img-compress [质量]` — 压缩图片\n\n"
        "📎 群聊中先发文件/图片，再 @我 发命令即可。"
    )


def _cmd_models() -> str:
    models = registry.list_models()
    lines = ["📋 **可用模型:**\n"]
    for m in models:
        status = "✅" if m["available"] else "❌"
        mid = m["model_id"]
        label = MODEL_LABELS.get(mid, mid)
        lines.append(f"• {status} `{mid}` — {label}")
    lines.append("\n使用 `/model <名称>` 切换。")
    return "\n".join(lines)


async def _cmd_model(open_id: str, text: str) -> str:
    parts = text.split(maxsplit=1)
    if len(parts) < 2:
        current = await get_session_model(open_id)
        label = MODEL_LABELS.get(current, current)
        return f"当前模型: **{label}** (`{current}`)\n使用 `/model <名称>` 切换"

    arg = parts[1].strip().lower()
    models = registry.list_models()
    matched = None
    for m in models:
        mid = m["model_id"].lower()
        if arg in mid or arg in MODEL_LABELS.get(m["model_id"], "").lower():
            matched = m
            break

    if not matched:
        available = ", ".join(m["model_id"] for m in models if m["available"])
        return f"未找到模型 `{arg}`。可用模型:\n{available}"

    if not matched["available"]:
        return f"模型 `{matched['model_id']}` 当前不可用"

    await set_session_model(open_id, matched["model_id"])
    label = MODEL_LABELS.get(matched["model_id"], matched["model_id"])
    return f"✅ 已切换到 **{label}** (`{matched['model_id']}`)"


async def _cmd_chat(open_id: str, text: str) -> str:
    """调用模型对话"""
    model = await get_session_model(open_id)

    try:
        adapter = registry.get(model)
    except ValueError:
        await set_session_model(open_id, DEFAULT_MODEL)
        adapter = registry.get(DEFAULT_MODEL)
        model = DEFAULT_MODEL

    if not adapter.model_info.available:
        return f"模型 `{model}` 不可用。使用 `/models` 查看可用模型。"

    try:
        result = await adapter.chat(messages=[{"role": "user", "content": text}])
    except Exception as e:
        logger.error(f"模型调用失败: model={model}, error={e}")
        err_msg = str(e).strip() or f"{type(e).__name__}（无详细错误信息）"
        return f"模型 `{model}` 调用失败。\n错误: {err_msg[:300]}"

    content = ""
    choices = result.get("choices", [])
    if choices:
        content = choices[0].get("message", {}).get("content", "")

    if not content:
        return "模型返回了空内容，请重试。"

    label = MODEL_LABELS.get(model, model)
    return f"{content}\n\n———\n📌 当前模型: **{label}**"


async def _cmd_read(open_id: str, chat_id: str, filename: str = "", question: str = "") -> str:
    """处理 /read 指令: 从群聊最近消息中找文件并分析

    用法:
      /read                    — 分析最近的文件
      /read report.pdf         — 分析指定文件
      /read report.pdf 总结一下 — 指定文件 + 自定义分析要求
    """
    if not chat_id:
        return (
            "📖 **读取文件**\n\n"
            "群聊中使用方法：\n"
            "1️⃣ 先把文件发到群里\n"
            "2️⃣ 再 @我 说「**/read**」分析最近文件\n"
            "   或「**/read 文件名**」指定文件\n\n"
            "💡 私聊直接发文件即可，无需指令。"
        )

    file_info = await fetch_recent_file(chat_id, lookback=20, filename=filename)
    if not file_info:
        hint = f"没有找到文件「{filename}」" if filename else "没有在最近消息里找到文件"
        return (
            f"{hint} 😅\n\n"
            "使用方法：\n"
            "1️⃣ 先把文件发到群里\n"
            "2️⃣ 再 @我 说「**/read**」或「**/read 文件名**」"
        )

    message_id = file_info["message_id"]
    file_key = file_info["file_key"]
    file_name = file_info.get("file_name", "unknown")

    logger.info(f"/read: 找到文件 {file_name} (key={file_key}, msg={message_id})")

    try:
        file_bytes = await download_resource(message_id, file_key, "file")
    except Exception as e:
        logger.error(f"/read 下载文件失败: {e}")
        return f"文件下载失败: {str(e)[:200]}"

    # 提取文本
    file_name_lower = file_name.lower()
    ext = file_name_lower[file_name_lower.rfind("."):] if "." in file_name_lower else ""

    if ext not in SUPPORTED_EXTENSIONS:
        return f"暂不支持 {ext} 格式的文件。\n支持: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"

    try:
        file_text = _extract_file_text(file_bytes, ext)
    except Exception as e:
        logger.error(f"/read 提取文本失败: {e}")
        return f"文件解析失败: {str(e)[:200]}"

    if not file_text:
        return "文件为空或无法提取文本内容。"

    # 截断过长文本
    max_chars = 12000
    if len(file_text) > max_chars:
        file_text = file_text[:max_chars] + f"\n\n...(文件过大，已截断，剩余 {len(file_text) - max_chars} 字符)"
    logger.info(f"提取文本 {len(file_text)} 字符，准备调用模型...")

    # 调用模型分析
    model = await get_session_model(open_id)
    label = MODEL_LABELS.get(model, model)

    try:
        adapter = registry.get(model)
    except ValueError:
        adapter = registry.get(DEFAULT_MODEL)
        model = DEFAULT_MODEL
        label = MODEL_LABELS.get(model, model)

    format_instruction = _get_format_instruction()
    if question:
        prompt = f"文件「{file_name}」内容：\n\n```\n{file_text}\n```\n\n{question}{format_instruction}"
    else:
        prompt = f"请分析以下文件内容（文件: {file_name}）：\n\n```\n{file_text}\n```\n\n请总结这个文件的核心内容。{format_instruction}"

    try:
        result = await adapter.chat(messages=[{"role": "user", "content": prompt}])
    except Exception as e:
        logger.error(f"/read 模型调用失败: type={type(e).__name__}, msg={e}")
        return f"文件分析失败: [{type(e).__name__}] {str(e)[:200] or '未知错误'}"

    content_text = ""
    choices = result.get("choices", [])
    if choices:
        content_text = choices[0].get("message", {}).get("content", "")

    if not content_text:
        return "模型未能分析文件内容。"

    return f"{content_text}\n\n———\n📁 **{file_name}** · 模型: **{label}**"


async def _cmd_search(open_id: str, query: str, message_id: str = "") -> str:
    """处理 /search 指令: 联网搜索 → 模型总结

    用法:
      /search 今天有什么新闻
      /search Python 最新特性
    """
    logger.info(f"/search: query={query[:100]}")

    # 0. 即时反馈 — 让用户知道机器人正在处理
    if message_id:
        await reply_text_message(message_id, f"🔍 正在搜索「{query[:50]}」...")

    # 1. 天气查询：自动识别 → 直接返回天气数据
    if is_weather_query(query):
        city_match = re.search(r'(北京|上海|广州|深圳|杭州|成都|武汉|南京|天津|重庆|西安|苏州|长沙|郑州|青岛|大连|厦门|福州|合肥|济南|沈阳|哈尔滨|昆明|贵阳|南宁|拉萨|乌鲁木齐|呼和浩特|银川|西宁|兰州|海口|三亚)', query)
        city = city_match.group(1) if city_match else "Beijing"
        weather = await get_weather(city)
        if weather:
            return f"{weather}\n\n———\n🔍 天气数据: wttr.in（实时）"
        # 天气 API 失败 → 回退到普通搜索
        logger.warning(f"天气 API 失败，回退到普通搜索")

    # 1. 搜索（附上当前日期，确保时效性）
    today = datetime.now().strftime("%Y年%m月%d日")
    search_query = query
    # 如果问题是"今天/今日"，追加日期到搜索词
    if any(word in query for word in ("今天", "今日", "现在", "当前", "最近", "最新")):
        search_query = f"{query} {today}"

    results = await search_web(search_query, max_results=8)
    if not results:
        return f"🔍 搜索「{query[:100]}」未找到结果，换个关键词试试？"

    # 2. 格式化搜索结果
    snippets = []
    for i, r in enumerate(results, 1):
        snippets.append(f"{i}. **{r['title']}**\n   链接: {r['url']}\n   摘要: {r['snippet']}")

    search_text = "\n\n".join(snippets)

    # 3. 构建 prompt（告知模型当前日期，帮助判断结果时效性）
    prompt = (
        f"当前日期：{today}\n\n"
        f"联网搜索结果：\n\n{search_text}\n\n"
        f"请根据以上搜索结果回答用户问题：{query}\n\n"
        f"要求：\n"
        f"1. 用飞书聊天可读的纯文本格式回复，合理分段\n"
        f"2. 引用信息时标注来源编号 [1] [2] 等\n"
        f"3. 从搜索摘要中尽可能提取、归纳、整理出有用的信息\n"
        f"4. 即使摘要信息不完整，也要把已有的整理出来，不要只说「搜索结果不够」就直接放弃\n"
        f"5. 可以补充常识性知识来完善回答，但注明哪些来自搜索结果、哪些来自常识"
    )

    # 4. 调用模型
    model = await get_session_model(open_id)
    try:
        adapter = registry.get(model)
    except ValueError:
        adapter = registry.get(DEFAULT_MODEL)
        model = DEFAULT_MODEL

    label = MODEL_LABELS.get(model, model)

    try:
        result = await adapter.chat(messages=[{"role": "user", "content": prompt}])
    except Exception as e:
        logger.error(f"/search 模型调用失败: {type(e).__name__}: {e}")
        return f"搜索完成但模型分析失败: [{type(e).__name__}] {str(e)[:200] or '未知错误'}"

    content = ""
    choices = result.get("choices", [])
    if choices:
        content = choices[0].get("message", {}).get("content", "")

    if not content:
        return "模型未能分析搜索结果。"

    return f"{content}\n\n———\n🔍 搜索结果 · 模型: **{label}**"


# ---- 媒体处理指令 ----


# 命令正则：/xxx 或 /xxx arg1 arg2
_MEDIA_CMD_RE = re.compile(
    r"^/(clip|compress|watermark|img-watermark|extract-audio|convert|resize|img-compress|concat|info)"
)

# 图片相关命令（用 Pillow 处理）
_IMAGE_CMDS = {"resize", "img-watermark", "img-compress"}


async def _cmd_media(open_id: str, chat_id: str, message_id: str, text: str) -> str:
    """处理媒体处理指令: 找最近文件 → 下载 → 处理 → 上传 → 回复

    支持命令:
      /clip <start> <end>        — 裁剪视频片段
      /compress [crf]            — 压缩视频/图片
      /watermark <text> [pos] [size]  — 视频加水印
      /img-watermark <text> [pos] [size] — 图片加水印
      /extract-audio [format]    — 提取音频
      /convert <format>          — 格式转换
      /resize <W> [H]            — 图片缩放
      /img-compress [quality]    — 图片压缩
      /concat                    — 拼接最近多个视频
      /info                      — 查看视频/图片信息
    """
    m = _MEDIA_CMD_RE.match(text)
    if not m:
        return "无法识别的媒体命令。输入 /help 查看帮助。"
    cmd = m.group(1)

    # 解析参数
    remaining = text[len(cmd) + 1:].strip()  # 去掉 /cmd
    args = remaining.split() if remaining else []

    # 拼接命令特殊处理 — 需要多个文件
    if cmd == "concat":
        return await _media_concat(chat_id, message_id)

    # 其余命令: 找最近文件
    if not chat_id:
        return "群聊中使用方法：\n1️⃣ 先把视频/图片发到群里\n2️⃣ 再 @我 说「**/{cmd}**」\n\n💡 私聊直接发文件，再用命令处理。"

    # 图片命令只找 image 类型，视频命令找 file 类型
    file_info = await _find_recent_media(chat_id, is_image=cmd in _IMAGE_CMDS)
    if not file_info:
        hint = "图片" if cmd in _IMAGE_CMDS else "视频/文件"
        return f"没有在最近消息里找到{hint} 😅\n\n请先把{hint}发到群里，再用 /{cmd} 处理。"

    # 即时反馈
    if message_id:
        await reply_text_message(message_id, f"⚙️ 正在处理「{file_info.get('file_name', 'unknown')}」...")

    # 下载文件
    try:
        file_bytes = await download_resource(
            file_info["message_id"], file_info["file_key"], file_info.get("resource_type", "file")
        )
    except Exception as e:
        logger.error(f"下载文件失败: {e}")
        return f"文件下载失败: {str(e)[:200]}"

    # 保存到临时文件
    fname = file_info.get("file_name", "temp")
    ext = fname[fname.rfind("."):] if "." in fname else ".mp4"
    input_path = f"/tmp/feishu_media/input_{int(time.time())}{ext}"
    os.makedirs("/tmp/feishu_media", exist_ok=True)
    with open(input_path, "wb") as f:
        f.write(file_bytes)

    logger.info(f"媒体文件已下载: {fname} ({len(file_bytes)} bytes) → {input_path}")

    # 路由到具体处理函数
    try:
        result_path = await _dispatch_media_cmd(cmd, args, input_path, fname)
    except Exception as e:
        logger.error(f"媒体处理失败: cmd={cmd}, error={e}", exc_info=True)
        _cleanup_temp(input_path)
        return f"处理失败: [{type(e).__name__}] {str(e)[:200]}"

    if not result_path:
        _cleanup_temp(input_path)
        return f"「{fname}」处理失败，请检查文件格式和参数是否正确。"

    # 上传结果到飞书
    is_image_output = os.path.splitext(result_path)[1].lower() in (".png", ".jpg", ".jpeg", ".webp", ".gif")
    try:
        if is_image_output:
            upload_result = await upload_image(result_path)
            img_key = upload_result.get("data", {}).get("image_key", "")
            if img_key:
                await reply_image_message(message_id, img_key)
            else:
                code = upload_result.get("code", -1)
                msg = upload_result.get("msg", "上传失败")
                _cleanup_temp(input_path, result_path)
                return f"图片上传失败: [{code}] {msg}"
        else:
            upload_result = await upload_file(result_path, "stream")
            file_key = upload_result.get("data", {}).get("file_key", "")
            if file_key:
                await reply_file_message(message_id, file_key)
            else:
                code = upload_result.get("code", -1)
                msg = upload_result.get("msg", "上传失败")
                _cleanup_temp(input_path, result_path)
                return f"文件上传失败: [{code}] {msg}"
    except Exception as e:
        logger.error(f"上传文件失败: {e}")
        _cleanup_temp(input_path, result_path)
        return f"文件上传失败: {str(e)[:200]}"

    # 生成说明文字
    desc = _media_result_desc(cmd, args, input_path, result_path)
    _cleanup_temp(input_path, result_path)
    return desc


def _cleanup_temp(*paths: str):
    """清理临时文件"""
    for p in paths:
        try:
            if p and os.path.exists(p):
                os.remove(p)
        except Exception:
            pass


async def _find_recent_media(chat_id: str, is_image: bool = False) -> dict | None:
    """从群聊最近消息中查找媒体文件（视频或图片）

    返回: {"message_id": "...", "file_key": "...", "file_name": "...", "resource_type": "file|image"}
    """
    from app.services.feishu_service import get_tenant_access_token, FEISHU_HOST

    token = await get_tenant_access_token()
    headers = {"Authorization": f"Bearer {token}"}
    url = (
        f"{FEISHU_HOST}/open-apis/im/v1/messages"
        f"?container_id_type=chat"
        f"&container_id={chat_id}"
        f"&page_size=30"
        f"&sort_type=ByCreateTimeDesc"
    )

    import httpx
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.get(url, headers=headers)
        data = resp.json()

    if data.get("code") != 0:
        logger.error(f"获取消息失败: {data}")
        return None

    items = data.get("data", {}).get("items", [])

    for msg in items:
        msg_type = msg.get("msg_type", "")
        body = msg.get("body", {})
        content_str = body.get("content", "{}")
        try:
            content = json.loads(content_str)
        except json.JSONDecodeError:
            continue

        if is_image and msg_type == "image":
            return {
                "message_id": msg.get("message_id", ""),
                "file_key": content.get("image_key", ""),
                "file_name": f"image_{msg.get('message_id', '')[:8]}.png",
                "resource_type": "image",
            }

        if not is_image and msg_type == "file":
            fname = content.get("file_name", body.get("title", "unknown"))
            # 视频文件
            if fname.lower().endswith((".mp4", ".mov", ".mkv", ".avi", ".webm", ".flv", ".wmv", ".m4v", ".mpg", ".mpeg", ".3gp")):
                return {
                    "message_id": msg.get("message_id", ""),
                    "file_key": content.get("file_key", ""),
                    "file_name": fname,
                    "resource_type": "file",
                }
            # 其他媒体文件 (gif等)
            if fname.lower().endswith((".gif",)):
                return {
                    "message_id": msg.get("message_id", ""),
                    "file_key": content.get("file_key", ""),
                    "file_name": fname,
                    "resource_type": "file",
                }

    # 如果 is_image=False 且没找到视频，也尝试找普通文件（可能是图片格式）
    if not is_image:
        for msg in items:
            msg_type = msg.get("msg_type", "")
            if msg_type == "image":
                body = msg.get("body", {})
                content_str = body.get("content", "{}")
                try:
                    content = json.loads(content_str)
                except json.JSONDecodeError:
                    continue
                return {
                    "message_id": msg.get("message_id", ""),
                    "file_key": content.get("image_key", ""),
                    "file_name": f"image_{msg.get('message_id', '')[:8]}.png",
                    "resource_type": "image",
                }

    return None


async def _media_concat(chat_id: str, message_id: str) -> str:
    """拼接最近多个视频"""
    from app.services.feishu_service import get_tenant_access_token, FEISHU_HOST
    import httpx

    token = await get_tenant_access_token()
    headers = {"Authorization": f"Bearer {token}"}
    url = (
        f"{FEISHU_HOST}/open-apis/im/v1/messages"
        f"?container_id_type=chat"
        f"&container_id={chat_id}"
        f"&page_size=50"
        f"&sort_type=ByCreateTimeAsc"  # 升序，保持发送顺序
    )

    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.get(url, headers=headers)
        data = resp.json()

    if data.get("code") != 0:
        return "获取消息列表失败。"

    items = data.get("data", {}).get("items", [])
    video_files = []
    for msg in items:
        if msg.get("msg_type") != "file":
            continue
        body = msg.get("body", {})
        content_str = body.get("content", "{}")
        try:
            content = json.loads(content_str)
        except json.JSONDecodeError:
            continue
        fname = content.get("file_name", "")
        if fname.lower().endswith((".mp4", ".mov", ".mkv", ".avi", ".webm", ".m4v")):
            video_files.append({
                "message_id": msg.get("message_id", ""),
                "file_key": content.get("file_key", ""),
                "file_name": fname,
            })

    if len(video_files) < 2:
        return f"需要至少 2 个视频才能拼接，只找到 {len(video_files)} 个。\n\n请按拼接顺序依次发送视频，再用 /concat。"

    if message_id and video_files:
        await reply_text_message(message_id, f"⚙️ 正在拼接 {len(video_files)} 个视频...")

    paths = []
    for vf in video_files:
        try:
            fb = await download_resource(vf["message_id"], vf["file_key"], "file")
        except Exception as e:
            for p in paths:
                _cleanup_temp(p)
            return f"下载「{vf['file_name']}」失败: {str(e)[:200]}"
        ext = vf["file_name"][vf["file_name"].rfind("."):] if "." in vf["file_name"] else ".mp4"
        tmp_path = f"/tmp/feishu_media/concat_{vf['message_id'][:8]}{ext}"
        os.makedirs("/tmp/feishu_media", exist_ok=True)
        with open(tmp_path, "wb") as f:
            f.write(fb)
        paths.append(tmp_path)

    result_path = await media.video_concat(paths)
    for p in paths:
        _cleanup_temp(p)

    if not result_path:
        return "视频拼接失败，请检查视频格式是否一致。"

    upload_result = await upload_file(result_path, "stream")
    file_key = upload_result.get("data", {}).get("file_key", "")
    if file_key and message_id:
        await reply_file_message(message_id, file_key)
    else:
        _cleanup_temp(result_path)
        return f"上传失败: {upload_result.get('msg', '未知错误')}"

    _cleanup_temp(result_path)
    return f"✅ {len(video_files)} 个视频拼接完成！"


_VALID_POSITIONS = {"top-left", "top-right", "bottom-left", "bottom-right", "top", "bottom", "center"}


def _parse_watermark_args(args: list[str]) -> tuple[str, str, int]:
    """解析水印参数

    格式: /watermark <文字> [位置] [字号]
    示例:
      /watermark 二愣子                        → 文字=二愣子, 位置=默认, 字号=0(自适应)
      /watermark 二愣子 top-left               → 指定位置
      /watermark 二愣子 bottom-right 48        → 指定位置+字号
      /watermark @公司名 center 36             → 指定位置+字号

    返回: (text, position, font_size)
    """
    text_parts = []
    position = "bottom-right"
    font_size = 0

    for a in args:
        if a.lower() in _VALID_POSITIONS:
            position = a.lower()
        elif a.isdigit():
            font_size = int(a)
        else:
            text_parts.append(a)

    return " ".join(text_parts), position, font_size


async def _dispatch_media_cmd(cmd: str, args: list[str], input_path: str, fname: str) -> str | None:
    """分发具体媒体命令"""
    ext = fname.lower()[fname.rfind("."):] if "." in fname else ""

    if cmd == "clip":
        if len(args) < 2:
            return None  # 参数不足，触发 usage 提示
        start, end = args[0], args[1]
        if not ext.endswith((".mp4", ".mov", ".mkv", ".avi", ".webm", ".m4v", ".flv", ".wmv", ".mpg", ".mpeg", ".3gp")):
            return None  # 非视频文件
        return await media.video_clip(input_path, start, end)

    elif cmd == "compress":
        # /compress [crf] [resolution]
        crf = 28
        resolution = ""
        for a in args:
            if a.isdigit():
                crf = int(a)
            elif ":" in a or a.isdigit():
                resolution = a
        # 判断是图片还是视频
        if ext in (".png", ".jpg", ".jpeg", ".webp"):
            return media.image_compress(input_path, quality=max(10, 100 - crf * 3))
        else:
            return await media.video_compress(input_path, crf=crf, resolution=resolution)

    elif cmd == "watermark":
        text, position, font_size = _parse_watermark_args(args)
        if not text:
            return None
        if ext in (".png", ".jpg", ".jpeg", ".webp"):
            return media.image_watermark(input_path, text, position, font_size)
        else:
            return await media.video_watermark(input_path, text, position, font_size)

    elif cmd == "img-watermark":
        text, position, font_size = _parse_watermark_args(args)
        if not text:
            return None
        return media.image_watermark(input_path, text, position, font_size)

    elif cmd == "extract-audio":
        fmt = args[0] if args else "mp3"
        return await media.video_extract_audio(input_path, fmt)

    elif cmd == "convert":
        if not args:
            return None
        return await media.video_convert(input_path, args[0])

    elif cmd == "resize":
        if not args:
            return None
        # /resize 800 或 /resize 800x600
        size = args[0]
        if "x" in size:
            w, h = size.split("x", 1)
            w, h = int(w), int(h)
        else:
            w, h = int(size), 0
        quality = int(args[1]) if len(args) > 1 else 85
        return media.image_resize(input_path, w, h, quality)

    elif cmd == "img-compress":
        quality = int(args[0]) if args else 75
        max_w = int(args[1]) if len(args) > 1 else 1920
        return media.image_compress(input_path, quality=quality, max_width=max_w)

    elif cmd == "info":
        info = await media.get_video_info(input_path)
        if not info:
            return None
        lines = [f"📹 **{fname}** 文件信息:\n"]
        if "duration" in info:
            lines.append(f"⏱ 时长: {info['duration']:.1f} 秒")
        if "width" in info:
            lines.append(f"📐 分辨率: {info['width']}x{info['height']} @ {info.get('fps', '?')}fps")
        if "size_mb" in info:
            lines.append(f"📦 大小: {info['size_mb']:.1f} MB")
        if "bitrate_kbps" in info:
            lines.append(f"📊 码率: {info['bitrate_kbps']} kbps")
        if "video_codec" in info:
            lines.append(f"🎞 编码: {info['video_codec']}")
        return "\n".join(lines)

    return None


def _media_result_desc(cmd: str, args: list[str], input_path: str, result_path: str) -> str:
    """生成处理结果描述"""
    orig_size = os.path.getsize(input_path) / 1024**2
    new_size = os.path.getsize(result_path) / 1024**2

    desc_map = {
        "clip": f"✅ 视频裁剪完成",
        "compress": f"✅ 压缩完成: {orig_size:.1f}MB → {new_size:.1f}MB",
        "watermark": "✅ 水印添加完成",
        "img-watermark": "✅ 图片水印添加完成",
        "extract-audio": "✅ 音频提取完成",
        "convert": f"✅ 格式转换完成 ({new_size:.1f}MB)",
        "resize": "✅ 图片缩放完成",
        "img-compress": f"✅ 图片压缩完成: {orig_size:.1f}MB → {new_size:.1f}MB",
        "info": "",  # info 直接返回文本
    }
    return desc_map.get(cmd, "✅ 处理完成")
